"""
Report generation for Phase 8.

Each `build_*` function returns (bytes, filename, content_type). Excel is built
with openpyxl; PDF with weasyprint (HTML -> PDF). `generate_report(run)` is the
entry point used by the async worker: it produces the file, attaches it to the
ReportRun and flips the status.
"""
import io
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import Q
from django.utils import timezone

from audit.models import AuditLog
from leaves.models import (
    Department, EnterpriseLeaveBalance, Leave, LeaveDayRecord, LeaveType,
    MonthlyLeaveSummary,
)
from .models import ReportRun, ReportType

User = get_user_model()

EXCEL_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_CT = "application/pdf"


# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------
def _year(params):
    return int(params.get("year") or timezone.now().year)


def _department_filter(qs, params, path="user__"):
    dept = params.get("department")
    if dept:
        qs = qs.filter(
            Q(**{f"{path}department__iexact": dept}) | Q(**{f"{path}department_ref__code__iexact": dept})
        )
    return qs


def _employee_label(user):
    return (user.get_full_name() or user.username) if user else "—"


def _num(v):
    return v if isinstance(v, Decimal) else Decimal(str(v or "0"))


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------
def _new_workbook():
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_sheet(wb, title, headers, rows):
    from openpyxl.styles import Font, PatternFill
    ws = wb.create_sheet(title=title[:31])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        width = max(len(str(header)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(header))
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = min(45, width + 2)
    ws.freeze_panes = "A2"
    return ws


def _workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# weasyprint helpers
# ---------------------------------------------------------------------------
_PDF_CSS = """
  @page { size: A4; margin: 1.6cm; @bottom-right { content: "Page " counter(page); font-size: 9px; color: #888; } }
  body { font-family: 'Helvetica', 'Arial', sans-serif; color: #1f2937; font-size: 11px; }
  h1 { font-size: 20px; color: #111827; margin: 0 0 4px; }
  h2 { font-size: 14px; color: #2563EB; border-bottom: 2px solid #e5e7eb; padding-bottom: 4px; margin-top: 22px; }
  .sub { color: #6b7280; font-size: 11px; margin-bottom: 16px; }
  table { width: 100%; border-collapse: collapse; margin: 8px 0; }
  th { background: #f3f4f6; text-align: left; padding: 6px 8px; font-size: 10px; text-transform: uppercase; color: #6b7280; }
  td { padding: 6px 8px; border-bottom: 1px solid #eee; }
  .bar-row { display: flex; align-items: center; gap: 8px; margin: 3px 0; }
  .bar-label { width: 130px; font-size: 10px; }
  .bar-track { flex: 1; background: #f3f4f6; border-radius: 3px; height: 14px; }
  .bar-fill { background: #2563EB; height: 14px; border-radius: 3px; }
  .kpi { display: inline-block; width: 30%; margin: 6px 1%; padding: 10px; background: #f9fafb; border-radius: 6px; }
  .kpi b { font-size: 20px; display: block; }
  .flag { color: #b91c1c; font-weight: 700; }
"""


def _render_pdf(inner_html, title):
    import weasyprint
    html = f"<html><head><meta charset='utf-8'><style>{_PDF_CSS}</style></head><body>{inner_html}</body></html>"
    return weasyprint.HTML(string=html).write_pdf()


def _bars(pairs, unit="days"):
    """HTML/CSS horizontal bar chart from [(label, value), ...]."""
    values = [float(v) for _, v in pairs] or [0]
    top = max(values) or 1
    rows = []
    for label, value in pairs:
        pct = (float(value) / top) * 100
        rows.append(
            f"<div class='bar-row'><div class='bar-label'>{label}</div>"
            f"<div class='bar-track'><div class='bar-fill' style='width:{pct:.1f}%'></div></div>"
            f"<div style='width:70px;text-align:right'>{value} {unit}</div></div>"
        )
    return "".join(rows) or "<p>No data.</p>"


# ===========================================================================
# 1. Employee Leave Register (Excel, 3 sheets)
# ===========================================================================
def build_employee_register(params):
    year = _year(params)
    users = _department_filter(User.objects.all().order_by("username"), params, path="")

    summary_rows = []
    for u in users:
        for b in EnterpriseLeaveBalance.objects.filter(user=u, year=year).select_related("leave_type"):
            summary_rows.append([
                _employee_label(u), u.email, getattr(u, "department", "") or "",
                b.leave_type.code, str(b.entitled_days), str(b.used_days),
                str(b.pending_days), str(b.available_days),
            ])

    detail_qs = LeaveDayRecord.objects.filter(year=year).select_related("user", "leave_type")
    detail_qs = _department_filter(detail_qs, params)
    detail_rows = [[
        _employee_label(r.user), str(r.date), r.leave_type.code, r.day_portion,
        r.status, "Yes" if r.is_weekend else "", "Yes" if r.is_holiday else "",
    ] for r in detail_qs.order_by("user__username", "date")]

    adj_rows = []
    for entry in AuditLog.objects.filter(changes__event="LEAVE_BALANCE_ADJUSTED").order_by("-created_at"):
        c = entry.changes
        adj_rows.append([
            str(entry.created_at.date()), _employee_label(entry.actor),
            c.get("leave_type", ""), c.get("delta", ""), c.get("reason", ""),
        ])

    wb = _new_workbook()
    _write_sheet(wb, "Summary", ["Employee", "Email", "Department", "Type", "Entitled", "Used", "Pending", "Available"], summary_rows)
    _write_sheet(wb, "Detail", ["Employee", "Date", "Type", "Portion", "Status", "Weekend", "Holiday"], detail_rows)
    _write_sheet(wb, "Adjustments", ["Date", "By", "Type", "Delta", "Reason"], adj_rows)
    return _workbook_bytes(wb), f"employee_register_{year}.xlsx", EXCEL_CT


# ===========================================================================
# 2. Monthly Attendance (Excel + PDF)
# ===========================================================================
def _attendance_rows(params):
    year = _year(params)
    qs = MonthlyLeaveSummary.objects.filter(year=year).select_related("user")
    qs = _department_filter(qs, params)
    if params.get("month"):
        qs = qs.filter(month=int(params["month"]))
    rows = []
    for s in qs.order_by("user__department", "user__username", "month"):
        rows.append({
            "department": getattr(s.user, "department", "") or "—",
            "employee": _employee_label(s.user),
            "month": s.month,
            "working_days": s.working_days,
            "leave_days": str(s.total_leave_days),
            "attendance": str(s.attendance_percentage),
        })
    return year, rows


def build_monthly_attendance_excel(params):
    year, rows = _attendance_rows(params)
    wb = _new_workbook()
    _write_sheet(
        wb, "Attendance",
        ["Department", "Employee", "Month", "Working Days", "Leave Days", "Attendance %"],
        [[r["department"], r["employee"], r["month"], r["working_days"], r["leave_days"], r["attendance"]] for r in rows],
    )
    return _workbook_bytes(wb), f"monthly_attendance_{year}.xlsx", EXCEL_CT


def build_monthly_attendance_pdf(params):
    year, rows = _attendance_rows(params)
    by_dept = {}
    for r in rows:
        by_dept.setdefault(r["department"], []).append(r)
    sections = []
    for dept, drows in sorted(by_dept.items()):
        body = "".join(
            f"<tr><td>{r['employee']}</td><td>{r['month']}</td><td>{r['working_days']}</td>"
            f"<td>{r['leave_days']}</td><td>{r['attendance']}%</td></tr>" for r in drows
        )
        sections.append(
            f"<h2>{dept}</h2><table><tr><th>Employee</th><th>Month</th><th>Working</th>"
            f"<th>Leave</th><th>Attendance</th></tr>{body}</table>"
        )
    inner = f"<h1>Monthly Attendance {year}</h1><div class='sub'>Grouped by department</div>{''.join(sections) or '<p>No data.</p>'}"
    return _render_pdf(inner, "Monthly Attendance"), f"monthly_attendance_{year}.pdf", PDF_CT


# ===========================================================================
# 3. Leave Utilization (PDF, charts)
# ===========================================================================
def build_leave_utilization(params):
    year = _year(params)
    records = _department_filter(
        LeaveDayRecord.objects.filter(year=year, status=LeaveDayRecord.Status.APPROVED, is_weekend=False, is_holiday=False),
        params,
    ).select_related("leave_type", "user")

    by_type, by_dept, by_month = {}, {}, {m: Decimal("0") for m in range(1, 13)}
    total = Decimal("0")
    for r in records:
        w = r.portion_days
        total += w
        by_type[r.leave_type.code] = by_type.get(r.leave_type.code, Decimal("0")) + w
        dep = getattr(r.user, "department", "") or "—"
        by_dept[dep] = by_dept.get(dep, Decimal("0")) + w
        by_month[r.month] += w

    prev = _department_filter(
        LeaveDayRecord.objects.filter(year=year - 1, status=LeaveDayRecord.Status.APPROVED, is_weekend=False, is_holiday=False),
        params,
    ).count()
    trend = "up" if records.count() > prev else ("down" if records.count() < prev else "flat")

    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    inner = f"""
      <h1>Leave Utilization {year}</h1>
      <div class='sub'>Executive summary</div>
      <div>
        <div class='kpi'><span>Total leave days</span><b>{total}</b></div>
        <div class='kpi'><span>Approved records</span><b>{records.count()}</b></div>
        <div class='kpi'><span>Trend vs {year - 1}</span><b>{trend.upper()}</b></div>
      </div>
      <h2>By leave type</h2>{_bars(sorted(by_type.items()))}
      <h2>By department</h2>{_bars(sorted(by_dept.items()))}
      <h2>By month</h2>{_bars([(month_labels[m - 1], by_month[m]) for m in range(1, 13)])}
    """
    return _render_pdf(inner, "Leave Utilization"), f"leave_utilization_{year}.pdf", PDF_CT


# ===========================================================================
# 4. Compliance (PDF)
# ===========================================================================
def build_compliance(params):
    year = _year(params)
    # High usage: used >= 80% of entitlement.
    high = []
    for b in EnterpriseLeaveBalance.objects.filter(year=year).select_related("user", "leave_type"):
        entitled = _num(b.entitled_days) + _num(b.carried_forward_days)
        if entitled > 0 and (_num(b.used_days) / entitled) >= Decimal("0.8"):
            pct = (_num(b.used_days) / entitled * 100).quantize(Decimal("0.1"))
            high.append((_employee_label(b.user), b.leave_type.code, f"{pct}%"))

    # Unusual pattern: frequent Monday/Friday approved leaves (>= 3).
    monfri = {}
    for r in LeaveDayRecord.objects.filter(year=year, status=LeaveDayRecord.Status.APPROVED).select_related("user"):
        if r.date.weekday() in (0, 4):
            monfri[r.user_id] = monfri.get(r.user_id, 0) + 1
    flagged_pattern = []
    for uid, count in monfri.items():
        if count >= 3:
            u = User.objects.filter(pk=uid).first()
            flagged_pattern.append((_employee_label(u), f"{count} Mon/Fri leaves"))

    # Document policy: approved leaves of types that require a document (we have
    # no document store on Leave, so these are surfaced for manual verification).
    doc_types = set(LeaveType.objects.filter(requires_document=True).values_list("code", flat=True))
    doc_flags = []
    for r in LeaveDayRecord.objects.filter(year=year, status=LeaveDayRecord.Status.APPROVED).select_related("user", "leave_type"):
        if r.leave_type.code in doc_types:
            doc_flags.append((_employee_label(r.user), r.leave_type.code, str(r.date)))

    def _tbl(headers, rows, empty="None."):
        if not rows:
            return f"<p>{empty}</p>"
        head = "".join(f"<th>{h}</th>" for h in headers)
        body = "".join("<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>" for row in rows)
        return f"<table><tr>{head}</tr>{body}</table>"

    inner = f"""
      <h1>Compliance Report {year}</h1>
      <div class='sub'>Employees and records requiring HR review</div>
      <h2 class='flag'>High leave usage (&ge; 80% of entitlement)</h2>{_tbl(['Employee', 'Type', 'Used %'], high)}
      <h2 class='flag'>Unusual patterns (frequent Mon/Fri)</h2>{_tbl(['Employee', 'Observation'], flagged_pattern)}
      <h2 class='flag'>Document-required leaves (verify certificates)</h2>{_tbl(['Employee', 'Type', 'Date'], doc_flags[:200])}
    """
    return _render_pdf(inner, "Compliance"), f"compliance_{year}.pdf", PDF_CT


# ===========================================================================
# 5. Audit Trail (Excel)
# ===========================================================================
def build_audit_trail(params):
    qs = AuditLog.objects.select_related("actor", "content_type").all()
    if params.get("date_from"):
        qs = qs.filter(created_at__date__gte=params["date_from"])
    if params.get("date_to"):
        qs = qs.filter(created_at__date__lte=params["date_to"])
    if params.get("actor_id"):
        qs = qs.filter(actor_id=params["actor_id"])
    if params.get("action"):
        qs = qs.filter(action=params["action"])

    rows = [[
        entry.created_at.isoformat(), _employee_label(entry.actor), entry.action,
        entry.object_repr, str(entry.changes), entry.ip_address or "",
    ] for entry in qs.order_by("-created_at")[:10000]]

    wb = _new_workbook()
    _write_sheet(wb, "Audit Trail", ["Timestamp", "Actor", "Action", "Target", "Changes", "IP"], rows)
    return _workbook_bytes(wb), "audit_trail.xlsx", EXCEL_CT


# ===========================================================================
# Dispatcher + async entry point
# ===========================================================================
def build_report_file(report_type, params):
    """Return (bytes, filename, content_type) for a report type + params."""
    fmt = (params.get("format") or "").lower()
    if report_type == ReportType.EMPLOYEE_REGISTER:
        return build_employee_register(params)
    if report_type == ReportType.MONTHLY_ATTENDANCE:
        return build_monthly_attendance_pdf(params) if fmt == "pdf" else build_monthly_attendance_excel(params)
    if report_type == ReportType.LEAVE_UTILIZATION:
        return build_leave_utilization(params)
    if report_type == ReportType.COMPLIANCE:
        return build_compliance(params)
    if report_type == ReportType.AUDIT_TRAIL:
        return build_audit_trail(params)
    raise ValueError(f"Unknown report type: {report_type}")


def generate_report(run: ReportRun):
    """Produce the file for a ReportRun and update its status. Safe to call in a
    background thread; all exceptions are captured onto the run."""
    run.status = ReportRun.Status.GENERATING
    run.save(update_fields=["status"])
    try:
        content, filename, _ct = build_report_file(run.report_type, run.params or {})
        run.file.save(filename, ContentFile(content), save=False)
        run.status = ReportRun.Status.READY
        run.completed_at = timezone.now()
        run.save(update_fields=["file", "status", "completed_at"])
    except Exception as exc:  # noqa: BLE001 - surface failure onto the run
        run.status = ReportRun.Status.FAILED
        run.error = f"{type(exc).__name__}: {exc}"
        run.completed_at = timezone.now()
        run.save(update_fields=["status", "error", "completed_at"])
    return run
